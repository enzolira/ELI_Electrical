from flask import render_template,redirect,session,request, flash, jsonify, url_for , send_file, make_response# type: ignore
from src import app
from src.models.user import User
from src.models.proyects import Proyect
from src.models.tgs import Tgs
from src.models.circuits import Circuit
from src.models.loads import Load
from src.models.tds import Tds
from src.models.total_tds import Total_tds
from src.controllers import report
from flask_bcrypt import Bcrypt # type: ignore
import math
import time
import pandas as pd # type: ignore
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side 
import numpy as np


#  -------------------------------------------------MAIN PAGE ---------------------------------------------------



@app.route('/loadbox/')
def loadbox():
    if 'user_id' not in session:
        return redirect('/logout')

    data ={'id': session['user_id']}

    user = User.get_by_id(data)
    proyects = Proyect.get_all_proyect_by_user_id(data)
    tgs = Proyect.get_all_tgs_by_proyect_id_and_user_id(data)
    wires= Proyect.get_all_wires()
    return render_template('loadbox.html', user=user, proyects=proyects, tgs=tgs, wires=wires)



#  -------------------------------------------------NEW PROYECTS ---------------------------------------------------



@app.route('/new_proyect', methods=['POST'])
def new_proyect():
    if 'user_id' not in session:
        return redirect('/logout')
    data = {
        'name': request.form['name'],
        'user_id': session['user_id']
    }

    Proyect.save(data)
    flash("Proyecto creado correctamente.","success_pro")
    return redirect('/loadbox')



# ------------------------------------------------ADD NEW GENERAL TABLE-------------------------------------------------



@app.route('/add_tgs', methods=['POST'])
def add_tgs():
    if 'user_id' not in session:
        return redirect('/logout')
    
    if not Tgs.validate_new_tg(request.form):
        return redirect('/loadbox')
    else:
        data = {
            'name': request.form['name'],
            'tag': request.form['tag'],
            'proyect_id': request.form['proyect_id']
        }
        Tgs.add_tgs(data)
        flash("Tablero general agregado correctamente", "tg_success")
    return redirect('/loadbox')


# ------------------------------------------------ADD NEW DISTRIBUCION TABLE--------------------------------------------



@app.route('/add_tds', methods=['POST'])
def add_tds():
    if 'user_id' not in session:
        return redirect('/logout')

    if not Tds.validate_new_td(request.form):
        return redirect('/loadbox')
    
    else: 
        data = {
            'name': request.form['name'],
            'tag': request.form['tag'],
            'tg_id': request.form['tg_id']
        }
        td_null = Tgs.get_tgs_circuit_tds_null({'tg_id': request.form['tg_id']})

        max_name = 0

        for td in td_null:
            print(td)
            name_value = int(td['name'])
            if name_value > max_name:
                max_name = name_value

        count1 = Tgs.count_tds({'tg_id': request.form['tg_id']})

        if max_name and count1:

            td = []
            td = Tds.get_td_by_id({'id':Tds.add_tds(data)})
            for qq in td:
                count = Tgs.get_tds_tgs_circuit({'tg_id':qq['tg_id']})
                print("Aca el if {count}")
                data2 = {
                    'name': count[0]['circuits_tg'],
                    'ref': qq['name'],
                    'tab_secondary': qq["tg_id"],
                    'td_id': qq["id"],
                    'method': request.form["method"],
                    'wires': request.form["type_isolation"],
                    'length_from_tg': request.form["length_from_tg"],
                }
                Total_tds.summary_tds(data2)

        elif not max_name or not count1:
            td = []
            td = Tds.get_td_by_id({'id':Tds.add_tds(data)})
            for qq in td:
                count = Tgs.get_tds_tgs_circuit({'tg_id':qq['tg_id']})
                print("Aca el elif")
                data2 = {
                    'name': count[0]['circuits_tg'],
                    'ref': qq['name'],
                    'tab_secondary': qq["tg_id"],
                    'td_id': qq["id"],
                    'method': request.form["method"],
                    'wires': request.form["type_isolation"],
                    'length_from_tg': request.form["length_from_tg"],
                }
                Total_tds.summary_tds(data2)
        flash("Tablero de distribución agregado correctamente", "td_success")
        return redirect('/loadbox')


# ------------------------------------------------ CREATE CIRCUITS-------------------------------------------------------


@app.route('/new_circuit', methods=['POST'])
def new_circuits():

    if 'user_id' not in session:
        return redirect('/logout')
    
    if not Circuit.validate_circuit(request.form):
        return redirect("/loadbox")
    
# --------------CALCULOS MONOFASICOS -------
    else:
        if float(request.form['single_voltage']) == 0.220:
            data_tg = {}
            data = {}
            data_tg['tg_id'] = request.form['tg_id']
            if 'td_id' in request.form and request.form['td_id'].isdigit():
                data_tg['td_id'] = request.form['td_id']
                rst = Circuit.all_r_s_t_single_voltage_tg_and_td(data_tg)
                rst2 = Total_tds.get_rst_total_tds_by_tg_id({'tab_secondary': data_tg['tg_id']})
            else:
                data_tg['td_id'] = '-Seleccione tablero de distribución-'
                rst = Circuit.all_r_s_t_single_voltage_tg({'tg_id':request.form["tg_id"]})
                rst2 = Total_tds.get_rst_total_tds_by_tg_id({'tab_secondary': data_tg['tg_id']})
            if int(data_tg['tg_id']) > 0 and data_tg['td_id'] == '-Seleccione tablero de distribución-':
                index_num = Tgs.get_tds_tgs_circuit({'tg_id': request.form['tg_id']})
                data = {
                    'name': index_num[0]['circuits_tg'] + 1,
                    'td_id': None
                }
            else:
                tds_circuit = Tds.get_all_circuits_by_td_id_and_tg_id(data_tg)
                data = {
                    'name': tds_circuit[0]['new_circuit_td'],
                    'td_id': request.form['td_id']
                }
            
            data.update({
                'nameloads': request.form['nameloads'],
                'ref': request.form['ref'],
                'method': request.form['method'],
                'single_voltage': request.form['single_voltage'],
                'total_length_ct': request.form['total_length_ct'],
                'tg_id' : request.form['tg_id'],
                'name_impedance': request.form["impedance"],
            })

            if data['name_impedance'] == "capacitance":
                data['total_fp'] = 0.95
                data['qty'] = request.form['qty']
                data['active_power'] = round(float(request.form['power']), 2)
                total_active_power = round((int(data['qty']) * float(data['active_power']))/1000,2)
                data['total_apparent_power'] = round((total_active_power) / float(data['total_fp']),2)
                data['total_reactive_power'] = round(-1 * ((data['total_apparent_power']) * math.sqrt(1 - data['total_fp']**2)),2)
            else:
                data['total_fp'] = 0.93
                data['qty'] = request.form['qty']
                data['active_power'] = round(float(request.form['power']), 2)
                total_active_power = round((int(data['qty']) * float(data['active_power']))/1000,2)
                data['total_apparent_power'] = round((total_active_power) / float(data['total_fp']),2)
                data['total_reactive_power'] = round(data['total_apparent_power'] * math.sqrt(1 - data['total_fp']**2),2)

            data['wires'] = request.form['type_isolation']
            data['type_circuit'] = request.form['type_circuit']
            circuit_id = Circuit.add_circuit(data)
            total_current = round(total_active_power/(float(request.form['single_voltage']) * float(data['total_fp'])),2)
            data1 = {'method':data['method'],'total_current':total_current}
            current_by_method= Proyect.current(data1)
            vp = (2 * 0.018 * float(total_current) * float(data['total_length_ct']))/float(current_by_method[0]['secction_mm2'])
            data2 = {}
            data2['circuit_id'] = circuit_id
            if vp < 4.5:
                current_by_method2 = Proyect.current(data1)
                data2['breakers'] = current_by_method2[0]['disyuntor']
                data2['elect_differencial'] = current_by_method2[0]['diferencial']
                data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                data2['current_by_method'] = current_by_method2[0][data['method']]
                Circuit.update_method(data2)
                Circuit.update_secctionmm2(data2)
                Circuit.update_breakers(data2)
                Circuit.update_elect_differencial(data2)
                def is_valid_current(current):
                    return (current is not None and
                            current[0]["R"] is not None and
                            current[0]["S"] is not None and
                            current[0]["T"] is not None)

                current_rst = Tgs.sum_all_current({'tg_id': request.form["tg_id"]})
                current_rst_td = Tds.sum_all_current_td_tg({'tg_id': data_tg['tg_id'], 'td_id': data_tg['td_id'].isdigit()})

                if is_valid_current(current_rst) or is_valid_current(current_rst_td):# Código si ambos son válidos


                # ----- calculating average from rst current ----------

                    avg = round((current_rst[0]["R"] + current_rst[0]["S"] + current_rst[0]["T"]) / 3,2)
                    unbalace_r = abs(round(current_rst[0]["R"] - avg,2))
                    unbalace_s = abs(round(current_rst[0]["S"] - avg,2))
                    unbalace_t = abs(round(current_rst[0]["T"] - avg,2))
                    porcentage_r = round((unbalace_r/avg)*100,2)
                    porcentage_s = round((unbalace_s/avg)*100,2)
                    porcentage_t = round((unbalace_t/avg)*100,2)
                    max_unbalance = max(porcentage_r, porcentage_s, porcentage_t)
                    print(f"{max_unbalance}%")
                    if max_unbalance < 10:       
#                   Diccionario que asocia etiquetas con los valores a comparar
                        updates = {'R': 'current_r', 'S': 'current_s', 'T': 'current_t'}

                        # Obtener claves y valores de rst y rst2
                        keys = ['R', 'S', 'T']
                        rst_values = rst[0]
                        rst2_values = rst2[0]

                        # Bucle para evaluar las relaciones
                        for i in range(len(keys)):
                            current_key = keys[i]

                            # Caso 1: Si todos los valores son iguales en rst y rst2, actualiza 'R'
                            if all(rst_values[k] == rst_values[keys[0]] for k in keys) and all(rst2_values[k] == rst2_values[keys[0]] for k in keys):
                                Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                                break  # Termina porque ya se actualizó correctamente

                            # Caso 2: Si el valor actual es mayor que el siguiente en ambas listas, actualiza el valor correspondiente
                            elif i < len(keys) - 1 and rst_values[current_key] > rst_values[keys[i + 1]] and rst2_values[current_key] > rst2_values[keys[i + 1]]:
                                Circuit.update_current_r({updates[keys[i + 1]]: total_current, 'circuit_id': circuit_id})
                                Total_tds.update_current_r_td({updates[keys[i + 1]]: total_current, 'td_id': data['td_id']})

                        print("aca desbalance bien, pero hay que revisar denuevo")

                    else:
                        print("aca si es mas de 10 agregar logica de mover la carga por entre fases")
                        print(f"ACA ESTOY ADENTRO")
                        if rst[0]['R'] == rst[0]['S'] == rst[0]['T']:
                            Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['R'] > rst[0]['S']:
                            Circuit.update_current_s({'current_s': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_s_td({'current_s': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['S'] > rst[0]['T']:
                            Circuit.update_current_t({'current_t': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_t_td({'current_t': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['T'] > rst[0]['R']:
                            Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                # ------------------------------------------------------

                else:
                    print(f"aca estoy cuando no hay rst")
                    # Bucle para evaluar las relaciones
                    updates = {
                        'R': 'current_r',
                        'S': 'current_s', 
                        'T': 'current_t'
                    }

                    rst_values = rst[0]  # Ej: {'R': 10, 'S': 8, 'T': 5}
                    rst2_values = rst2[0]  # Ej: {'R': 9, 'S': 7, 'T': 4}

                    # --- Caso 1: Todos los valores son iguales en ambas lecturas ---
                    if all(val == rst_values['R'] for val in rst_values.values()) and \
                    all(val == rst2_values['R'] for val in rst2_values.values()):
                        
                        # Actualizar solo 'R' (fase por defecto)
                        column = 'current_r'
                        update_data = {column: total_current, 'circuit_id': circuit_id}
                        Circuit.update_current_r(update_data)
                        
                        if data['td_id'] and str(data['td_id']).isdigit():
                            Total_tds.update_current_r_td({column: total_current, 'td_id': data['td_id']})

                    # --- Caso 2: Encontrar la fase con mayor diferencia desbalanceada ---
                    else:
                        # Calcular diferencias entre fases (rst)
                        diff_rst = {
                            'R': rst_values['R'] - rst_values['S'],
                            'S': rst_values['S'] - rst_values['T'],
                            'T': rst_values['T'] - rst_values['R']  # Cierre del triángulo
                        }
                        
                        # Calcular diferencias entre fases (rst2)
                        diff_rst2 = {
                            'R': rst2_values['R'] - rst2_values['S'],
                            'S': rst2_values['S'] - rst2_values['T'],
                            'T': rst2_values['T'] - rst2_values['R']
                        }
                        
                        # Fase a actualizar: Mayor diferencia en ambas lecturas
                        phase_to_update = max(
                            ['R', 'S', 'T'],
                            key=lambda phase: diff_rst[phase] + diff_rst2[phase]
                        )
                        
                        # Obtener columna dinámica (current_r, current_s, current_t)
                        column = updates[phase_to_update]
                        
                        # Actualizar en Circuit
                        Circuit.update_current_r({column: total_current, 'circuit_id': circuit_id})
                        
                        # Actualizar en Total_tds (si td_id es válido)
                        if data['td_id'] and str(data['td_id']).isdigit():
                            Total_tds.update_current_r_td({column: total_current, 'td_id': data['td_id']})

                data3 = { 'vp':round((2*0.018*float(data['total_length_ct'])*float(total_current))/float(data2['secctionmm2']),2), 'circuit_id':circuit_id}
                Circuit.update_vp(data3)
                data4 = {'nameloads':data['nameloads'],'qty': data['qty'], 'active_power':round(float(data['active_power']),2), 'total_active_power':total_active_power, 'total_apparent_power':float(data['total_apparent_power']), 'total_reactive_power':float(data['total_reactive_power']),'fp': data['total_fp'], 'impedance': data['name_impedance'],'length': request.form['total_length_ct'],'voltage': data['single_voltage'] , 'total_current': total_current,'circuit_id': circuit_id }
                Load.save(data4)
                Circuit.updated_loads({'circuit_id':circuit_id})
                if data['method'] == 'd1' or data['method'] == 'd2':
                    conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c3'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
                else:
                    conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c3'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
                index = Tgs.total_name_tg({'tg_id': data_tg['tg_id']})
                index2 = Total_tds.name_total_tds({'tab_secondary': data_tg['tg_id']})
                ordered_circuits = sorted(index, key=lambda x: int(x['name']))
                
                new_names = list(range(1, len(ordered_circuits) + 1))
                updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
                
                for circuit_id, new_name in updates.items():
                    Tgs.update_name({'id': circuit_id, 'name': new_name})

                last_number = new_names[-1] if new_names else 0
                new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
                updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}
                for circuit_id, new_name in updates_index2.items():
                    Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
                flash("Circuito agregado correctamente", "circuit_success")
                time.sleep(1)
                return redirect('/loadbox')
            else:
                allcurrent_by_method = Circuit.vp_real(data1)
                for all_current in allcurrent_by_method:
                    print(all_current)
                    if float(2 * 0.018 * float(total_current) * float(data['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                        data2['secctionmm2'] = all_current['secction_mm2']
                        data2['current_by_method'] = all_current['method']
                        break

                current_by_method2 = Proyect.current(data1)
                data2['breakers'] = current_by_method2[0]['disyuntor']
                data2['elect_differencial'] = current_by_method2[0]['diferencial']
                Circuit.update_method(data2)
                Circuit.update_secctionmm2(data2)
                Circuit.update_breakers(data2)
                Circuit.update_elect_differencial(data2)
                current_rst = Tgs.sum_all_current({'tg_id':data_tg['td_id']})
                print(current_rst)
                if current_rst is not None and current_rst[0]["R"] is not None and current_rst[0]["S"] is not None and current_rst[0]["T"] is not None:

                # ----- calculating average from rst current ----------

                    avg = round((current_rst[0]["R"] + current_rst[0]["S"] + current_rst[0]["T"]) / 3,2)
                    unbalace_r = abs(round(current_rst[0]["R"] - avg,2))
                    unbalace_s = abs(round(current_rst[0]["S"] - avg,2))
                    unbalace_t = abs(round(current_rst[0]["T"] - avg,2))
                    porcentage_r = round((unbalace_r/avg)*100,2)
                    porcentage_s = round((unbalace_s/avg)*100,2)
                    porcentage_t = round((unbalace_t/avg)*100,2)
                    max_unbalance = max(porcentage_r, porcentage_s, porcentage_t)
                    print(f"{max_unbalance}%")
                    if max_unbalance < 10:       
                        if rst[0]['R'] == rst[0]['S'] == rst[0]['T']:
                            Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['R'] > rst[0]['S']:
                            Circuit.update_current_s({'current_s': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_s_td({'current_s': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['S'] > rst[0]['T']:
                            Circuit.update_current_t({'current_t': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_t_td({'current_t': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['T'] > rst[0]['R']:
                            Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        print("aca desbalance bien, pero hay que revisar denuevo")

                    else:
                        print("aca si es mas de 10 agregar logica de mover la carga por entre fases")
                        print(f"aca estoy cuando hay rst")
                        if rst[0]['R'] == rst[0]['S'] == rst[0]['T']:
                            Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['R'] > rst[0]['S']:
                            Circuit.update_current_s({'current_s': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_s_td({'current_s': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['S'] > rst[0]['T']:
                            Circuit.update_current_t({'current_t': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_t_td({'current_t': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                        if rst[0]['T'] > rst[0]['R']:
                            Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                            if data['td_id'] is not None and str(data['td_id']).isdigit():
                                Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                            else:
                                pass
                # ------------------------------------------------------

                else:
                    print(f"aca estoy cuando no hay rst")
                    if rst[0]['R'] == rst[0]['S'] == rst[0]['T']:
                        Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                        if data['td_id'] is not None and str(data['td_id']).isdigit():
                            Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                        else:
                            pass
                    if rst[0]['R'] > rst[0]['S']:
                        Circuit.update_current_s({'current_s': total_current, 'circuit_id': circuit_id})
                        if data['td_id'] is not None and str(data['td_id']).isdigit():
                            Total_tds.update_current_s_td({'current_s': total_current, 'td_id': data['td_id']})
                        else:
                            pass
                    if rst[0]['S'] > rst[0]['T']:
                        Circuit.update_current_t({'current_t': total_current, 'circuit_id': circuit_id})
                        if data['td_id'] is not None and str(data['td_id']).isdigit():
                            Total_tds.update_current_t_td({'current_t': total_current, 'td_id': data['td_id']})
                        else:
                            pass
                    if rst[0]['T'] > rst[0]['R']:
                        Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                        if data['td_id'] is not None and str(data['td_id']).isdigit():
                            Total_tds.update_current_r_td({'current_r': total_current, 'td_id': data['td_id']})
                        else:
                            pass

                data3 = { 'vp':round((2*0.018*float(data['total_length_ct'])*float(total_current))/float(data2['secctionmm2']),2), 'circuit_id':circuit_id}
                Circuit.update_vp(data3)
                print(data['name_impedance'])
                data4 = {'nameloads':data['nameloads'],'qty': data['qty'], 'active_power':round(float(data['active_power']),2), 'total_active_power':total_active_power, 'total_apparent_power':float(data['total_apparent_power']), 'total_reactive_power':float(data['total_reactive_power']),'fp': data['total_fp'], 'impedance': data['name_impedance'], 'voltage': data['single_voltage'] , 'length': request.form['total_length_ct'], 'total_current': total_current,'circuit_id': circuit_id}
                Load.save(data4)
                Circuit.updated_loads({'circuit_id':circuit_id})
                if data['method'] == 'd1' or data['method'] == 'd2':
                    conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c3'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
                else:
                    conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c3'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
            index = Tgs.total_name_tg({'tg_id': data_tg['tg_id']})
            index2 = Total_tds.name_total_tds({'tab_secondary': data_tg['tg_id']})
            ordered_circuits = sorted(index, key=lambda x: int(x['name']))
            
            new_names = list(range(1, len(ordered_circuits) + 1))
            updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
            
            for circuit_id, new_name in updates.items():
                Tgs.update_name({'id': circuit_id, 'name': new_name})

            last_number = new_names[-1] if new_names else 0
            new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
            updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}
            for circuit_id, new_name in updates_index2.items():
                Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
            print('funciona por fin 220')
            flash("Circuito agregado correctamente", "circuit_success")
            time.sleep(1)
            return redirect('/loadbox')


    # ------------------ CALCULOS TRIFASICOS --------------------
        else:
            data_tg = {}
            data = {}
            data_tg['tg_id'] = request.form['tg_id']
            if 'td_id' in request.form and request.form['td_id']:
                data_tg['td_id'] = request.form['td_id']
            else:
                data_tg["td_id"] = '-Seleccione tablero de distribución-'

            if int(data_tg['tg_id']) > 0 and data_tg['td_id'] == '-Seleccione tablero de distribución-':
                index_num = Tgs.get_tds_tgs_circuit({'tg_id': request.form['tg_id']})
                data = {
                    'name': index_num[0]['circuits_tg'] + 1,
                    'td_id': None
                }
                print(data['name'])
            else:
                tds_circuit = Tds.get_all_circuits_by_td_id_and_tg_id(data_tg)
                data = {
                    'name': tds_circuit[0]['new_circuit_td'],
                    'td_id': request.form['td_id']
                }
                print(data['name'])

            data.update({
                'nameloads': request.form['nameloads'],
                'ref': request.form['ref'],
                'method': request.form['method'],
                'single_voltage': request.form['single_voltage'],
                'total_length_ct': request.form['total_length_ct'],
                'name_impedance': request.form['impedance'],
                'total_fp': request.form['fp'],
                'tg_id' : request.form['tg_id']
            })
            if data['name_impedance'] == "capacitance":
                data['qty'] = request.form['qty']
                data['active_power'] = round(float(request.form['power']), 2)
                total_active_power = round((int(data['qty']) * float(data['active_power']))/1000,2)
                data['total_apparent_power'] = round((total_active_power) / float(request.form['fp']),2)
                data['total_reactive_power'] = round(-1 * ((data['total_apparent_power']) * math.sqrt(1 - float(request.form['fp'])**2)),2)
            else:
                data['qty'] = request.form['qty']
                data['active_power'] = round(float(request.form['power']), 2)
                total_active_power = round((int(data['qty']) * float(data['active_power']))/1000,2)
                data['total_apparent_power'] = round((total_active_power) / float(data['total_fp']),2)
                data['total_reactive_power'] = round(data['total_apparent_power'] * math.sqrt(1 - float(request.form['fp'])),2)

            data['wires'] = request.form['type_isolation']
            data['type_circuit'] = request.form['type_circuit']
            circuit_id = Circuit.add_circuit(data)
            total_current = round(total_active_power/(float(request.form['single_voltage']) * math.sqrt(3) * float(request.form['fp'])),2)
            data1 = {'method':data['method'],'total_current':total_current}
            current_by_method= Proyect.current_tri(data1)
            vp = (0.018 * float(total_current) * float(data['total_length_ct']))/float(current_by_method[0]['secction_mm2'])
            data2 = {}
            data2['circuit_id'] = circuit_id
            if vp < 4.5:
                current_by_method2 = Proyect.current_tri(data1)
                data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                data2['current_by_method'] = current_by_method2[0][data['method']]
                data2['breakers'] = current_by_method2[0]['disyuntor']
                data2['elect_differencial'] = current_by_method2[0]['diferencial']
                Circuit.update_method(data2)
                Circuit.update_secctionmm2(data2)
                Circuit.update_breakers(data2)
                Circuit.update_elect_differencial(data2)
                Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                Circuit.update_current_s({'current_s': total_current, 'circuit_id': circuit_id})
                Circuit.update_current_t({'current_t': total_current, 'circuit_id': circuit_id})
                data3 = { 'vp':round((0.018*float(data['total_length_ct'])*float(total_current))/float(data2['secctionmm2']),2), 'circuit_id':circuit_id}
                Circuit.update_vp(data3)
                data4 = {'nameloads':data['nameloads'],'qty': data['qty'], 'active_power':round(float(data['active_power']),2), 'total_active_power':total_active_power, 'total_apparent_power':float(data['total_apparent_power']), 'total_reactive_power':float(data['total_reactive_power']),'fp':float(data['total_fp']), 'impedance': data['name_impedance'], 'voltage': data['single_voltage'] , 'length': request.form['total_length_ct'], 'total_current': total_current,'circuit_id': circuit_id}
                Load.save(data4)
                Circuit.updated_loads({'circuit_id':circuit_id})
                if data['method'] == 'd1' or data['method'] == 'd2':
                    print(data['method'])
                    conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c5'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
                else:
                    conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c5'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
                    index = Tgs.total_name_tg({'tg_id': data_tg['tg_id']})
                index2 = Total_tds.name_total_tds({'tab_secondary': data_tg['tg_id']})
                ordered_circuits = sorted(index, key=lambda x: int(x['name']))
                
                new_names = list(range(1, len(ordered_circuits) + 1))
                updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
                
                for circuit_id, new_name in updates.items():
                    Tgs.update_name({'id': circuit_id, 'name': new_name})

                last_number = new_names[-1] if new_names else 0
                new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
                updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}
                for circuit_id, new_name in updates_index2.items():
                    Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
                print('funciona por fin 380')
                flash("Circuito agregado correctamente", "circuit_success")
                time.sleep(1)
                return redirect('/loadbox')
            else:
                allcurrent_by_method = Circuit.vp_real(data1)
                for all_current in allcurrent_by_method:
                    if float(0.018 *float(total_current) * float(data['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                        data2['current_by_method'] =  all_current['method']
                        data2['secctionmm2'] = all_current['secction_mm2']
                        break

                current_by_method2 = Proyect.current_tri(data1)
                data2['current_by_method'] = current_by_method2[0][data['method']]
                data2['breakers'] = current_by_method2[0]['disyuntor']
                data2['elect_differencial'] = current_by_method2[0]['diferencial']
                Circuit.update_method(data2)
                Circuit.update_secctionmm2(data2)
                Circuit.update_breakers(data2)
                Circuit.update_elect_differencial(data2)
                Circuit.update_current_r({'current_r': total_current, 'circuit_id': circuit_id})
                Circuit.update_current_s({'current_s': total_current, 'circuit_id': circuit_id})
                Circuit.update_current_t({'current_t': total_current, 'circuit_id': circuit_id})
                data3 = { 'vp':round((0.018*float(data['total_length_ct'])*float(total_current))/float(data2['secctionmm2']),2), 'circuit_id':circuit_id}
                Circuit.update_vp(data3)
                data4 = {'nameloads':data['nameloads'],'qty': data['qty'], 'active_power':round(float(data['active_power']),2), 'total_active_power':total_active_power, 'total_apparent_power':float(data['total_apparent_power']), 'total_reactive_power':float(data['total_reactive_power']),'fp':float(data['total_fp']), 'impedance': data['name_impedance'], 'voltage': data['single_voltage'] , 'length': request.form['total_length_ct'], 'total_current': total_current,'circuit_id': circuit_id}
                Load.save(data4)
                Circuit.updated_loads({'circuit_id':circuit_id})
                if data['method'] == 'd1' or data['method'] == 'd2':
                    conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c5'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
                else:
                    conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c5'], 'circuit_id':circuit_id}
                    Circuit.update_conduit(data5)
            index = Tgs.total_name_tg({'tg_id': data_tg['tg_id']})
            index2 = Total_tds.name_total_tds({'tab_secondary': data_tg['tg_id']})
            ordered_circuits = sorted(index, key=lambda x: int(x['name']))
            
            new_names = list(range(1, len(ordered_circuits) + 1))
            updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
            
            for circuit_id, new_name in updates.items():
                Tgs.update_name({'id': circuit_id, 'name': new_name})

            last_number = new_names[-1] if new_names else 0
            new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
            updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}
            for circuit_id, new_name in updates_index2.items():
                Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
            print('funciona por fin 380')
            flash("Circuito agregado correctamente", "circuit_success")
            time.sleep(1)
            return redirect('/loadbox')


# ---------------------------------------------- AJAX ID CIRCUITS----------------------------------------------------------------


@app.route('/api/tgs', methods=['POST'])
def get_tgs():
    proyect = request.form['proyect']
    tgs1 = Tgs.get_tgs_by_project({'proyect_id': proyect})
    return jsonify(tgs1)


@app.route('/api/tds', methods=['POST'])
def get_tds():
    tg_id = request.form['tgs']
    circuit_tg = Circuit.get_all_circuits_by_tg_id({'tg_id':tg_id})
    tgs = Tds.get_all_tds_by_tg_id({'tg_id':tg_id})
    data2 = {}
    total_tds = []
    info_tds = []
    test_1 = []
    test_2 = []
    test_tds = Total_tds.get_all_total_tds()

    for ss in test_tds:
        if ss['total_current_ct']:
            print(ss['single_voltage'])
            # ---------------- VP MONOFASICO ---------------
            if float(ss['single_voltage']) == 0.220:
                data1 = {'method':ss['method'],'total_current':ss['total_current_ct']}
                current_by_method= Proyect.current(data1)
                vp = round(2 * 0.018 * float(ss['total_current_ct']) * float(ss['length_from_tg'])/float(current_by_method[0]['secction_mm2']),2)
                print(vp)
                data2 = {}
                data2['td_id'] = ss['td_id']
                if vp < 4.5:
                    current_by_method2 = Proyect.current(data1)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                    data2['current_by_method'] = current_by_method2[0][ss['method']]
                    Total_tds.update_method_total_td(data2)
                    Total_tds.update_secctionmm2_total_td(data2)
                    Total_tds.update_breakers_total_td(data2)
                    Total_tds.update_elect_differencial_total_td(data2)
                    Total_tds.update_vp_td({'vp': vp , 'td_id': ss['td_id']})
                    if ss['method'] == 'd1' or ss['method'] == 'd2':
                        conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                        print(conduit)
                        datos = {'conduit': conduit[0]['c3'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos)
                    else:
                        conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                        datos2 = {'conduit': conduit[0]['c3'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos2)
                else:
                    allcurrent_by_method = Circuit.vp_real(data1)
                    for all_current in allcurrent_by_method:
                        if float(2 * 0.018 * float(ss['total_current_ct']) * float(ss['length_from_tg']))/float(all_current['secction_mm2']) < 4.5:
                            data2['secctionmm2'] = all_current['secction_mm2']
                            data2['current_by_method'] = all_current['method']
                            break

                    current_by_method2 = Proyect.current(data1)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    Total_tds.update_method_total_td(data2)
                    Total_tds.update_secctionmm2_total_td(data2)
                    Total_tds.update_breakers_total_td(data2)
                    Total_tds.update_elect_differencial_total_td(data2)
                    vp_new = round(2 * 0.018 * float(ss['total_current_ct']) * float(ss['length_from_tg'])/float(data2['secctionmm2']),2)
                    Total_tds.update_vp_td({'vp': vp_new , 'td_id': ss['td_id']})
                    if ss['method'] == 'd1' or ss['method'] == 'd2':
                        conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                        datos3 = {'conduit': conduit[0]['c3'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos3)
                    else:
                        conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                        datos4 = {'conduit': conduit[0]['c3'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos4)
                print('actualizado total_tds 220')

            # ---------------- VP TRIFASICO ---------------
            else:
                data1 = {'method':ss['method'],'total_current':ss['total_current_ct']}
                current_by_method= Proyect.current_tri(data1)
                vp = round(0.018 * float(ss['total_current_ct']) * float(ss['length_from_tg'])/float(current_by_method[0]['secction_mm2']),2)
                data2 = {}
                data2['td_id'] = ss['td_id']
                if vp < 4.5:
                    current_by_method2 = Proyect.current_tri(data1)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                    data2['current_by_method'] = current_by_method2[0][ss['method']]
                    Total_tds.update_method_total_td(data2)
                    Total_tds.update_secctionmm2_total_td(data2)
                    Total_tds.update_breakers_total_td(data2)
                    Total_tds.update_elect_differencial_total_td(data2)
                    Total_tds.update_current_r_td({'current_r':ss['total_current_ct'], 'td_id':ss['td_id']})
                    Total_tds.update_current_s_td({'current_s':ss['total_current_ct'], 'td_id':ss['td_id']})
                    Total_tds.update_current_t_td({'current_t':ss['total_current_ct'], 'td_id':ss['td_id']})
                    Total_tds.update_vp_td({'vp': vp , 'td_id': ss['td_id']})
                    if ss['method'] == 'd1' or ss['method'] == 'd2':
                        conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                        datos5 = {'conduit': conduit[0]['c5'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos5)
                    else:
                        conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                        datos6 = {'conduit': conduit[0]['c5'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos6)
                else:
                    allcurrent_by_method = Circuit.vp_real(data1)
                    for all_current in allcurrent_by_method:
                        print(all_current)
                        if float(0.018 * float(ss['total_current_ct']) * float(ss['length_from_tg']))/float(all_current['secction_mm2']) < 4.5:
                            data2['secctionmm2'] = all_current['secction_mm2']
                            data2['current_by_method'] = all_current['method']
                            break

                    current_by_method2 = Proyect.current_tri(data1)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    Total_tds.update_method_total_td(data2)
                    Total_tds.update_secctionmm2_total_td(data2)
                    Total_tds.update_breakers_total_td(data2)
                    Total_tds.update_elect_differencial_total_td(data2)
                    Total_tds.update_current_r_td({'current_r':ss['total_current_ct'], 'td_id':ss['td_id']})
                    Total_tds.update_current_s_td({'current_s':ss['total_current_ct'], 'td_id':ss['td_id']})
                    Total_tds.update_current_t_td({'current_t':ss['total_current_ct'], 'td_id':ss['td_id']})
                    vp_new = round(0.018 * float(ss['total_current_ct']) * float(ss['length_from_tg'])/float(current_by_method2[0]['secction_mm2']),2)
                    Total_tds.update_vp_td({'vp': vp_new , 'td_id': ss['td_id']})
                    if ss['method'] == 'd1' or ss['method'] == 'd2':
                        conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                        datos7 = {'conduit': conduit[0]['c5'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos7)
                    else:
                        conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                        datos8 = {'conduit': conduit[0]['c5'], 'td_id':ss['td_id']}
                        Total_tds.update_conduit_total_td(datos8)
                print('actualizado total_tds 380')
        else:
            pass
        test_1.append(ss)

    for jj in tgs:
        data2['td_id'] = jj['id']
        data2['tg_id'] = jj['tg_id']
        info_tds.append(Circuit.detail_total_circuits_by_td_id(data2))

    print("-------------")
    for ll in info_tds:
        for rr in ll:
            test_2.append(rr)
            if rr:
                all_impedance = Circuit.all_name_impedance({'td_id':rr['td_id']})
                # ---UPDATE IMPEDANCE -----
                count_cap = 0
                count_ind = 0
                for uu in all_impedance:
                    if uu["name_impedance"] == "capacitance":
                        count_cap += 1
                    else:
                        count_ind += 1
                    if count_cap >= count_ind:
                        if rr['all_fp'] >= 0.95 and (rr['total_reactive_power_ct'] < 0 or rr['total_reactive_power_ct'] > 0):
                            print("no hay multa cap")
                        elif 0.95 > rr['all_fp'] >= 0.93 and rr['total_reactive_power_ct'] > 0:
                            print("hay multa cap 1")
                        else:
                            print("hay multa cap revisar armonicos 2")
                        Total_tds.update_td_impedance({'td_impedance': 'capacitance', 'td_id': rr['td_id']})
                    else:
                        if 0.95 > rr['all_fp'] >= 0.93 and rr['total_reactive_power_ct'] < 0:
                            print("hay multa cap 3")
                            Total_tds.update_td_impedance({'td_impedance': 'capacitance', 'td_id': rr['td_id']})
                        else:
                            print("no hay multa inductiva")
                        Total_tds.update_td_impedance({'td_impedance': 'inductance', 'td_id': rr['td_id']})
            else:
                pass
                # ---------- FIN UPDATE -------------------
    print("-------------")

    for dict_1, dict_2 in zip(test_1, test_2):
        total_center_1 = dict_1.get('total_center')
        total_center_2 = dict_2.get('total_center')


        if total_center_1 == total_center_2:
            print(f"Los valores de 'total_center' son iguales para '{dict_1['ref']}' y '{dict_2['ref']}': {total_center_1}, {dict_1['name']}")

        else:
            print(f"Valor en test_1 ({dict_1['ref']}): {total_center_1}")
            print(f"Valor en test_2 ({dict_2['ref']}): {total_center_2}")
            if total_center_1 != total_center_2:
                print("Valor diferente en test_2:")
                Total_tds.update_total_tds(dict_2)
            else:
                print("Valor diferente en test_1:")
                Total_tds.update_total_tds(dict_1)


    result = Total_tds.get_all_total_tds_by_tg_id({'tab_secondary':tg_id})
    if result:
        total_tds.append(result)
    print("---------------")

    return jsonify(tgs, circuit_tg, total_tds)



@app.route('/api/circuits_td', methods=['POST'])
def get_all_circuits_by_tds():

    tgs_values = request.form.getlist('tgs[]')
    tds_values = request.form.getlist('tds[]')

    data = {
        'tg_id': tgs_values,
        'td_id': tds_values
    }
    circuit_td = {}
    if data['td_id'] and data['tg_id']:
        circuit_td = Circuit.get_all_circuit_and_tds_by_tg(data)
    else:
        pass
    return jsonify(circuit_td)

# --------------------- FROM AJAX ABOUT DETAIL CIRCUITS BY TDS AND TGS -----------------------------------------------------------

@app.route('/api/detail', methods=['POST'])
def detai_circuit_tds():
    circuit_id = request.form['circuit']
    circuitos = Circuit.detail_circuit_and_loads_by_id({'circuit_id': circuit_id})
    return jsonify(circuitos)


@app.route('/add_load', methods=['POST'])
def add_loads():
    circuit = Circuit.detail_circuit_and_loads_by_id({'circuit_id': request.form['circuit_id']})

    if not Load.validate_add_load(request.form):
        return redirect("/loadbox")
# ------- ADD LOADS IN CIRCUITS MONOPHASE -----------------
    else:
        if float(circuit[0]['single_voltage']) == 0.220:
            load = {
                'nameloads': request.form['nameloads'],
                'qty': request.form['qty'],
                'circuit_id': request.form['circuit_id'],
                'active_power': request.form['power'],
                'impedance': request.form['impedance'],
                'length': request.form['total_length_ct'],
                'voltage': request.form['voltage']
            }
            print(f" la tension de la carga es: {request.form['voltage']}")
            if load["impedance"] == "capacitance":
                load['fp'] = 0.95
                total_active_power = round(float((load['qty'])) * float(load['active_power'])/1000,2)
                load['total_active_power'] = total_active_power
                load['total_apparent_power'] = round(total_active_power / float(load["fp"]),2)
                load['total_reactive_power'] = round(-1 * (float(load["total_apparent_power"]) * math.sqrt(1 - load['fp']**2)),2)
            else:
                load['fp'] = 0.93
                total_active_power = round(float((load['qty'])) * float(load['active_power'])/1000,2)
                load['total_active_power'] = total_active_power
                load['total_apparent_power'] = round(total_active_power / float(load["fp"]),2)
                load['total_reactive_power'] = round(float(load["total_apparent_power"]) * math.sqrt(1 - load['fp']**2),2)

            total_current = round(total_active_power/(float(load['voltage']) * float(load['fp'])),2)
            load['total_current'] = total_current
            Load.save(load)
            Circuit.updated_loads({'circuit_id': request.form['circuit_id']})

    # ---------------------- CALCULO IMPEDANCIA Y FP ----------------------------------------

            new_total_fp = Circuit.new_total_fp({'id': request.form['circuit_id']})
            print(f" el valor de total_fp es {new_total_fp[0]['power_factor']}")
            Circuit.update_total_fp({'id': request.form['circuit_id'], 'new_total_fp':new_total_fp[0]['power_factor']})
            all_impedance = Load.all_impedance({'circuit_id': request.form['circuit_id']})
            count_cap = 0
            count_ind = 0
            for uu in all_impedance:
                if uu["impedance"] == "capacitance":
                    count_cap += 1
                else:
                    count_ind += 1

            if count_cap >= count_ind:
                if new_total_fp[0]['power_factor'] >= 0.95:
                    if load['total_reactive_power'] < 0:
                        print("no hay multa cap 1")
                    else:
                        print("no hay multa cap 2")
                elif 0.95 > new_total_fp[0]['power_factor'] >= 0.93:
                    if load['total_reactive_power'] > 0:
                        print("hay multa cap 1")
                    else:
                        print("hay multa cap revisar armonicos 2")
                Circuit.update_name_impedance({'name_impedance': 'capacitance', 'id': request.form['circuit_id']})
            else:
                if new_total_fp[0]['power_factor'] >= 0.93:
                    if load['total_reactive_power'] < 0:
                        print("hay multa cap 3")
                    else:
                        print("no hay multa cap aun que hay mas ind")
                    Circuit.update_name_impedance({'name_impedance': 'capacitance', 'id': request.form['circuit_id']})
                else:
                    print("hay multa cap revisar armonicos 2")
                    Circuit.update_name_impedance({'name_impedance': 'inductance', 'id': request.form['circuit_id']})


    # ---------------------------------------------------------------------------------------
            Newtotal_current_ct = Circuit.detail_circuit_and_loads_by_id({'circuit_id': request.form['circuit_id']})
            data = {
                'method':circuit[0]['method'],
                'total_current': Newtotal_current_ct[0]['total_current_ct']
            }
            vp = round((2 * 0.018 * float(Newtotal_current_ct[0]['total_current_ct']) * float(Newtotal_current_ct[0]['total_length_ct']))/float(Newtotal_current_ct[0]['secctionmm2']),2)
            data2 = {}
            data2['circuit_id'] = request.form['circuit_id']
            if vp < 4.5:
                current_by_method2 = Proyect.current(data)
                data2['breakers'] = current_by_method2[0]['disyuntor']
                data2['elect_differencial'] = current_by_method2[0]['diferencial']
                data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                data2['current_by_method'] = current_by_method2[0][data['method']]
                Circuit.update_method(data2)
                Circuit.update_secctionmm2(data2)
                Circuit.update_breakers(data2)
                Circuit.update_elect_differencial(data2)
                data3 = { 'vp':round((2*0.018*float(Newtotal_current_ct[0]['total_length_ct'])*float(Newtotal_current_ct[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit_id']}
                Circuit.update_vp(data3)
                Circuit.updated_loads({'circuit_id': request.form['circuit_id']})
                if Newtotal_current_ct[0]['method'] == 'd1' or Newtotal_current_ct[0]['method'] == 'd2':
                    conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c3'], 'circuit_id':request.form['circuit_id']}
                    Circuit.update_conduit(data5)
                    flash("Carga agregada correctamente", "load_success")
                    return redirect('/loadbox/')
                else:
                    conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                    data5 = {'conduit': conduit[0]['c3'], 'circuit_id':request.form['circuit_id']}
                    Circuit.update_conduit(data5)
                print('funciona por fin 220')
                flash("Carga agregada correctamente", "load_success")
                return redirect('/loadbox')
            else:
                allcurrent_by_method = Circuit.vp_real(data)
                for all_current in allcurrent_by_method:
                    print(all_current)
                    if float(2 * 0.018 * float(Newtotal_current_ct[0]['total_current_ct']) * float(Newtotal_current_ct[0]['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                        data2['secctionmm2'] = all_current['secction_mm2']
                        data2['current_by_method'] = all_current['method']
                        break

            current_by_method2 = Proyect.current(data)
            data2['breakers'] = current_by_method2[0]['disyuntor']
            data2['elect_differencial'] = current_by_method2[0]['diferencial']
            Circuit.update_method(data2)
            Circuit.update_secctionmm2(data2)
            Circuit.update_breakers(data2)
            Circuit.update_elect_differencial(data2)
            data3 = { 'vp':round((2*0.018*float(Newtotal_current_ct[0]['total_length_ct'])*float(Newtotal_current_ct[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit_id']}
            Circuit.update_vp(data3)
            Circuit.updated_loads({'circuit_id': request.form['circuit_id']})
            if Newtotal_current_ct[0]['method'] == 'd1' or Newtotal_current_ct[0]['method'] == 'd2':
                conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                data5 = {'conduit': conduit[0]['c3'], 'circuit_id':request.form['circuit_id']}
                Circuit.update_conduit(data5)
                flash("Carga Monofásica agregada correctamente", "load_success")
                return redirect('/loadbox/')
            else:
                conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                data5 = {'conduit': conduit[0]['c3'], 'circuit_id':request.form['circuit_id']}
                Circuit.update_conduit(data5)
            print('funciona por fin 220')
            flash("Carga agregada correctamente", "load_success")
            return redirect('/loadbox')

    # ------- ADD LOADS IN CIRCUITS TRIPHASE -----------------
        else:
            print("aca estamos")
    # ------ 380/220-------
            voltage = request.form["voltage2"]
            print(voltage)
            load = {
            'nameloads': request.form['nameloads'],
            'qty': request.form['qty'],
            'circuit_id': request.form['circuit_id'],
            'active_power': request.form['power'],
            'impedance': request.form['impedance2'],
            'length': request.form['total_length_ct'],
            'voltage': voltage
            }

            if voltage == "0.220":
                if load['impedance'] == "capacitance":
                    load['fp'] = 0.95
                    total_active_power = round((int(load['qty']) * float(load['active_power']))/1000,2)
                    load['total_active_power'] = total_active_power
                    load['total_apparent_power'] = round((total_active_power) / float(load['fp']),2)
                    load['total_reactive_power'] = round(-1 * ((load['total_apparent_power']) * math.sqrt(1 - float(load['fp'])**2)),2)
                else:
                    load['fp'] = 0.93
                    total_active_power = round((int(load['qty']) * float(load['active_power']))/1000,2)
                    load['total_active_power'] = total_active_power
                    load['total_apparent_power'] = round((total_active_power) / float(load['fp']),2)
                    load['total_reactive_power'] = round(load['total_apparent_power'] * math.sqrt(1 - float(load['fp'])),2)

                total_current = round(total_active_power/(math.sqrt(3) * float(float(voltage)) * float(load['fp'])),2)
                load['total_current'] = total_current
                Load.save(load)
                Circuit.updated_loads({'circuit_id': request.form['circuit_id']})

    # ---------------------- CALCULO IMPEDANCIA Y FP ----------------------------------------

                new_total_fp = Circuit.new_total_fp({'id': request.form['circuit_id']})
                print(f" el valor de total_fp es {new_total_fp[0]['power_factor']}")
                Circuit.update_total_fp({'id': request.form['circuit_id'], 'new_total_fp':new_total_fp[0]['power_factor']})
                all_impedance = Load.all_impedance({'circuit_id': request.form['circuit_id']})
                count_cap = 0
                count_ind = 0
                for uu in all_impedance:
                    if uu["impedance"] == "capacitance":
                        count_cap += 1
                    else:
                        count_ind += 1

                if count_cap >= count_ind:
                    if new_total_fp[0]['power_factor'] >= 0.95:
                        if load['total_reactive_power'] < 0:
                            print("no hay multa cap 1")
                        else:
                            print("no hay multa cap 2")
                    elif 0.95 > new_total_fp[0]['power_factor'] >= 0.93:
                        if load['total_reactive_power'] > 0:
                            print("hay multa cap 1")
                        else:
                            print("hay multa cap revisar armonicos 2")
                    Circuit.update_name_impedance({'name_impedance': 'capacitance', 'id': request.form['circuit_id']})
                else:
                    if new_total_fp[0]['power_factor'] >= 0.93:
                        if load['total_reactive_power'] < 0:
                            print("hay multa cap 3")
                        else:
                            print("no hay multa cap aun que hay mas ind")
                        Circuit.update_name_impedance({'name_impedance': 'capacitance', 'id': request.form['circuit_id']})
                    else:
                        print("hay multa cap revisar armonicos 2")
                        Circuit.update_name_impedance({'name_impedance': 'inductance', 'id': request.form['circuit_id']})

    # ---------------------------------------------------------------------------------------

                Newtotal_current_ct = Circuit.detail_circuit_and_loads_by_id({'circuit_id': request.form['circuit_id']})
                data2 = {}
                data2['circuit_id'] = request.form['circuit_id']
                data = {
                    'method':circuit[0]['method'],
                    'total_current': Newtotal_current_ct[0]['total_current_ct']
                }
                vp = round((0.018 * float(Newtotal_current_ct[0]['total_current_ct']) * float(Newtotal_current_ct[0]['total_length_ct']))/float(Newtotal_current_ct[0]['secctionmm2']),2)
                if vp < 4.5:
                    current_by_method2 = Proyect.current_tri(data)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                    data2['current_by_method'] = current_by_method2[0][data['method']]
                    Circuit.update_method(data2)
                    Circuit.update_secctionmm2(data2)
                    Circuit.update_breakers(data2)
                    Circuit.update_elect_differencial(data2)
                    Circuit.update_current_r({'current_r': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_s({'current_s': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_t({'current_t': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    data3 = { 'vp':round((0.018*float(request.form['total_length_ct'])*float(Newtotal_current_ct[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit_id']}
                    Circuit.update_vp(data3)
                    Circuit.updated_loads({'circuit_id': request.form['circuit_id']})
                    if data['method'] == 'd1' or data['method'] == 'd2':
                        conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                        flash("Carga agregada correctamente", "load_success")
                        return redirect('/loadbox/')
                    else:
                        conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                    print('funciona por fin 380/220')
                    flash("Carga agregada correctamente", "load_success")
                    return redirect('/loadbox/')
                else:
                    allcurrent_by_method = Circuit.vp_real(data)
                    for all_current in allcurrent_by_method:
                        print(all_current)
                        if float(0.018 * float(Newtotal_current_ct[0]['total_current_ct']) * float(Newtotal_current_ct[0]['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                            data2['secctionmm2'] = all_current['secction_mm2']
                            data2['current_by_method'] = all_current['method']
                            break

                    current_by_method2 = Proyect.current_tri(data)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    Circuit.update_method(data2)
                    Circuit.update_secctionmm2(data2)
                    Circuit.update_breakers(data2)
                    Circuit.update_elect_differencial(data2)
                    Circuit.update_current_r({'current_r': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_s({'current_s': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_t({'current_t': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    data3 = { 'vp':round((0.018*float(Newtotal_current_ct[0]['total_length_ct'])*float(Newtotal_current_ct[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit_id']}
                    Circuit.update_vp(data3)
                    Circuit.updated_loads({'circuit_id': request.form['circuit_id']})
                    if data['method'] == 'd1' or data['method'] == 'd2':
                        conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                        flash("Carga agregada correctamente", "load_success")
                        return redirect('/loadbox/')
                    else:
                        conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                    print('funciona por fin 380/220')
                    flash("Carga agregada correctamente", "load_success")
                    return redirect('/loadbox/')

    # ------- 380/380--------
            else:

                load['fp'] = request.form["fp3"]
                total_active_power = round((int(load['qty']) * float(load['active_power']))/1000,2)
                load['total_active_power'] = total_active_power
                load['total_apparent_power'] = round((total_active_power) / float(load['fp']),2)

                if load['impedance'] == "capacitance":
                    load['total_reactive_power'] = round(-1 * ((load['total_apparent_power']) * math.sqrt(1 - float(load['fp'])**2)),2)

                else:
                    load['total_reactive_power'] = round((load['total_apparent_power']) * math.sqrt(1 - float(load['fp'])**2),2)

                total_current = round(total_active_power/(math.sqrt(3) * float(float(voltage)) * float(load['fp'])),2)
                load['total_current'] = total_current
                Load.save(load)
                Circuit.updated_loads({'circuit_id': request.form['circuit_id']})

    # ---------------------- CALCULO IMPEDANCIA Y FP ----------------------------------------

                new_total_fp = Circuit.new_total_fp({'id': request.form['circuit_id']})
                print(f" el valor de total_fp es {new_total_fp[0]['power_factor']}")
                Circuit.update_total_fp({'id': request.form['circuit_id'], 'new_total_fp':new_total_fp[0]['power_factor']})
                all_impedance = Load.all_impedance({'circuit_id': request.form['circuit_id']})
                count_cap = 0
                count_ind = 0
                for uu in all_impedance:
                    if uu["impedance"] == "capacitance":
                        count_cap += 1
                    else:
                        count_ind += 1

                if count_cap >= count_ind:
                    if new_total_fp[0]['power_factor'] >= 0.95:
                        if load['total_reactive_power'] < 0:
                            print("no hay multa cap 1")
                        else:
                            print("no hay multa cap 2")
                    elif 0.95 > new_total_fp[0]['power_factor'] >= 0.93:
                        if load['total_reactive_power'] > 0:
                            print("hay multa cap 1")
                        else:
                            print("hay multa cap revisar armonicos 2")
                    Circuit.update_name_impedance({'name_impedance': 'capacitance', 'id': request.form['circuit_id']})
                else:
                    if new_total_fp[0]['power_factor'] >= 0.93:
                        if load['total_reactive_power'] < 0:
                            print("hay multa cap 3")
                        else:
                            print("no hay multa cap aun que hay mas ind")
                        Circuit.update_name_impedance({'name_impedance': 'capacitance', 'id': request.form['circuit_id']})
                    else:
                        print("hay multa cap revisar armonicos 2")
                        Circuit.update_name_impedance({'name_impedance': 'inductance', 'id': request.form['circuit_id']})
    # ---------------------------------------------------------------------------------------

                Newtotal_current_ct = Circuit.detail_circuit_and_loads_by_id({'circuit_id': request.form['circuit_id']})
                data2 = {}
                data2['circuit_id'] = request.form['circuit_id']
                data = {
                    'method':circuit[0]['method'],
                    'total_current': Newtotal_current_ct[0]['total_current_ct']
                }
                vp = round((0.018 * float(Newtotal_current_ct[0]['total_current_ct']) * float(Newtotal_current_ct[0]['total_length_ct']))/float(Newtotal_current_ct[0]['secctionmm2']),2)
                if vp < 4.5:
                    current_by_method2 = Proyect.current_tri(data)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    data2['secctionmm2'] = current_by_method2[0]['secction_mm2']
                    data2['current_by_method'] = current_by_method2[0][data['method']]
                    Circuit.update_method(data2)
                    Circuit.update_secctionmm2(data2)
                    Circuit.update_breakers(data2)
                    Circuit.update_elect_differencial(data2)
                    Circuit.update_current_r({'current_r': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_s({'current_s': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_t({'current_t': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    data3 = { 'vp':round((0.018*float(request.form['total_length_ct'])*float(Newtotal_current_ct[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit_id']}
                    Circuit.update_vp(data3)
                    Circuit.updated_loads({'circuit_id': request.form['circuit_id']})
                    if data['method'] == 'd1' or data['method'] == 'd2':
                        conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                        flash("Carga agregada correctamente", "load_success")
                        return redirect('/loadbox/')
                    else:
                        conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                    print('funciona por fin 380/380')
                    flash("Carga agregada correctamente", "load_success")
                    return redirect('/loadbox/')
                else:
                    allcurrent_by_method = Circuit.vp_real(data)
                    for all_current in allcurrent_by_method:
                        print(all_current)
                        if float(0.018 * float(Newtotal_current_ct[0]['total_current_ct']) * float(Newtotal_current_ct[0]['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                            data2['secctionmm2'] = all_current['secction_mm2']
                            data2['current_by_method'] = all_current['method']
                            break

                    current_by_method2 = Proyect.current_tri(data)
                    data2['breakers'] = current_by_method2[0]['disyuntor']
                    data2['elect_differencial'] = current_by_method2[0]['diferencial']
                    Circuit.update_method(data2)
                    Circuit.update_secctionmm2(data2)
                    Circuit.update_breakers(data2)
                    Circuit.update_elect_differencial(data2)
                    Circuit.update_current_r({'current_r': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_s({'current_s': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    Circuit.update_current_t({'current_t': Newtotal_current_ct[0]['total_current_ct'], 'circuit_id': data2['circuit_id']})
                    data3 = { 'vp':round((0.018*float(Newtotal_current_ct[0]['total_length_ct'])*float(Newtotal_current_ct[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit_id']}
                    Circuit.update_vp(data3)
                    Circuit.updated_loads({'circuit_id': request.form['circuit_id']})
                    if data['method'] == 'd1' or data['method'] == 'd2':
                        conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                        flash("Carga agregada correctamente", "load_success")
                        return redirect('/loadbox/')
                    else:
                        conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                        data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit_id']}
                        Circuit.update_conduit(data5)
                    print('funciona por fin 380/380')
                    flash("Carga agregada correctamente", "load_success")
                    return redirect('/loadbox/')


# --------------------- FROM AJAX ABOUT EDIT AND DELETE CIRCUITS BY TDS AND TGS -----------------------------------------------------------


@app.route('/api/delete/load', methods=["POST"])
def delete_load():
    Load.delete({'load_id':request.form['load']})
    print('borrado')
    Circuit.updated_loads({'circuit_id':request.form['circuit']})
    totalCircuit = Circuit.detail_circuit_by_id({'circuit_id':request.form['circuit']})
    print(totalCircuit[0]['total_center'])
    print(totalCircuit)

    if totalCircuit[0]['total_center'] is not None:

# ---------------------- CALCULO IMPEDANCIA Y FP ----------------------------------------
        new_total_fp = Circuit.new_total_fp({'id': request.form['circuit']})
        print(f" el valor de total_fp es {new_total_fp[0]['power_factor']}")
        Circuit.update_total_fp({'id': request.form['circuit'], 'new_total_fp':new_total_fp[0]['power_factor']})
# ---------------------------------------------------------------------------------------

        circuitv2 = Circuit.detail_circuit_and_loads_by_id({'circuit_id':request.form['circuit']})
        if float(circuitv2[0]['single_voltage']) == 0.220:
            data2 = {}
            data2['circuit_id'] = request.form['circuit']
            data = {
                'method': circuitv2[0]['method'],
                'total_current': circuitv2[0]['total_current_ct']
            }
            allcurrent_by_method = Circuit.vp_real(data)
            for all_current in allcurrent_by_method:
                if float(2 * 0.018 * float(circuitv2[0]['total_current_ct']) * float(circuitv2[0]['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                    data2['secctionmm2'] = all_current['secction_mm2']
                    data2['current_by_method'] = all_current['method']
                    break

            current_by_method2 = Proyect.current(data)
            data2['breakers'] = current_by_method2[0]['disyuntor']
            data2['elect_differencial'] = current_by_method2[0]['diferencial']
            Circuit.update_method(data2)
            Circuit.update_secctionmm2(data2)
            Circuit.update_breakers(data2)
            Circuit.update_elect_differencial(data2)
            data3 = { 'vp':round((2*0.018*float(circuitv2[0]['total_length_ct'])*float(circuitv2[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id': request.form['circuit']}
            Circuit.update_vp(data3)
            Circuit.updated_loads({'circuit_id': request.form['circuit']})
            if circuitv2[0]['method'] == 'd1' or circuitv2[0]['method'] == 'd2':
                conduit = Circuit.conduit_mono_subte({'secctionmm2':data2['secctionmm2']})
                data5 = {'conduit': conduit[0]['c3'], 'circuit_id': request.form['circuit']}
                Circuit.update_conduit(data5)
            else:
                conduit = Circuit.conduit_mono_normal({'secctionmm2':data2['secctionmm2']})
                data5 = {'conduit': conduit[0]['c3'], 'circuit_id': request.form['circuit']}
                Circuit.update_conduit(data5)
            print('funciona por fin 220')
            return redirect('/loadbox/')

        # -------- trifasico ----------
        else:
            data2 = {}
            data2['circuit_id'] = request.form['circuit']
            data = {
                'method': circuitv2[0]['method'],
                'total_current': circuitv2[0]['total_current_ct']
            }
            allcurrent_by_method = Circuit.vp_real(data)
            for all_current in allcurrent_by_method:
                if float(0.018 * float(circuitv2[0]['total_current_ct']) * float(circuitv2[0]['total_length_ct']))/float((all_current['secction_mm2'])) < 4.5:
                    data2['secctionmm2'] = all_current['secction_mm2']
                    data2['current_by_method'] = all_current['method']
                    break

            current_by_method2 = Proyect.current_tri(data)
            data2['breakers'] = current_by_method2[0]['disyuntor']
            data2['elect_differencial'] = current_by_method2[0]['diferencial']
            Circuit.update_method(data2)
            Circuit.update_secctionmm2(data2)
            Circuit.update_breakers(data2)
            Circuit.update_elect_differencial(data2)
            data3 = { 'vp':round((0.018*float(circuitv2[0]['total_length_ct'])*float(circuitv2[0]['total_current_ct']))/float(data2['secctionmm2']),2), 'circuit_id':request.form['circuit']}
            Circuit.update_vp(data3)
            Circuit.updated_loads({'circuit_id': request.form['circuit']})
            if circuitv2[0]['method'] == 'd1' or circuitv2[0]['method'] == 'd2':
                conduit = Circuit.conduit_tri_subte({'secctionmm2':data2['secctionmm2']})
                data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit']}
                Circuit.update_conduit(data5)
            else:
                conduit = Circuit.conduit_tri_normal({'secctionmm2':data2['secctionmm2']})
                data5 = {'conduit': conduit[0]['c5'], 'circuit_id':request.form['circuit']}
                Circuit.update_conduit(data5)
            print('funciona por fin 380')
        return redirect('/loadbox/')
    else:
        print(request.form)
        tg = Tgs.tg_id_by_circuit({'circuit_id': request.form['circuit']})
        td = Tds.td_id_by_circuits({'circuit_id': request.form['circuit']})
        if not td[0]['td_id']:
            print(tg)
            print(td)
            Circuit.delete_circuit_by_id({'circuit_id': request.form['circuit']})
            index = Tgs.total_name_tg({'tg_id': tg[0]['tg_id']})
            index2 = Total_tds.name_total_tds({'tab_secondary': tg[0]['tg_id']})
            ordered_circuits = sorted(index, key=lambda x: int(x['name']))
            
            new_names = list(range(1, len(ordered_circuits) + 1))
            updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
            
            for circuit_id, new_name in updates.items():
                Tgs.update_name({'id': circuit_id, 'name': new_name})

            last_number = new_names[-1] if new_names else 0
            new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
            updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}
            for circuit_id, new_name in updates_index2.items():
                Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
                
            print("borrado del lado de loads 1")      
            return redirect('/loadbox/')
        else:
            print(tg)
            print(td)
            Circuit.delete_circuit_by_id({'circuit_id': request.form['circuit']})
            index = Tds.total_name_td({'td_id': td[0]['td_id']})
            ordered_circuits = sorted(index, key=lambda x: int(x['name']))
            
            new_names = list(range(1, len(ordered_circuits) + 1))
            updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
            
            for circuit_id, new_name in updates.items():
                Tgs.update_name({'id': circuit_id, 'name': new_name})

            print("borrado del lado de loads 2")      
        return redirect('/loadbox/')

@app.route('/api/delete/circuit', methods=['POST'])
def delete_circuit():
    print(request.form)
    tg = Tgs.tg_id_by_circuit({'circuit_id': request.form['circuitv2']})
    td = Tds.td_id_by_circuits({'circuit_id': request.form['circuitv2']})
    if not td[0]['td_id']:
        print(tg)
        print(td)
        Load.delete_load_by_circuit_id({'circuit_id': request.form['circuitv2']})
        Circuit.delete_circuit_by_id({'circuit_id': request.form['circuitv2']})
        index = Tgs.total_name_tg({'tg_id': tg[0]['tg_id']})
        index2 = Total_tds.name_total_tds({'tab_secondary': tg[0]['tg_id']})
        ordered_circuits = sorted(index, key=lambda x: int(x['id']))
        
        new_names = list(range(1, len(ordered_circuits) + 1))
        updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
        
        for circuit_id, new_name in updates.items():
            Tgs.update_name({'id': circuit_id, 'name': new_name})

        last_number = new_names[-1] if new_names else 0
        new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
        updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}
        for circuit_id, new_name in updates_index2.items():
            Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
        
        return redirect('/loadbox/')
    else:
        print(tg)
        print(td)
        Load.delete_load_by_circuit_id({'circuit_id': request.form['circuitv2']})
        Circuit.delete_circuit_by_id({'circuit_id': request.form['circuitv2']})
        index = Tds.total_name_td({'td_id': td[0]['td_id']})
        ordered_circuits = sorted(index, key=lambda x: int(x['id']))
        
        new_names = list(range(1, len(ordered_circuits) + 1))
        updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}
        
        for circuit_id, new_name in updates.items():
            Tgs.update_name({'id': circuit_id, 'name': new_name})
        
        return redirect('/loadbox/')


@app.route('/api/delete/tgs', methods=['POST'])
def delete_tgs():
    Tgs.delete_load_by_tgId({'tg_id':request.form['tgs_delete']})
    Tgs.delete_circuits_by_tgId({'tg_id':request.form['tgs_delete']})
    Tgs.delete_total_tgs_by_tgId({'tab_secondary':request.form['tgs_delete']})
    Tgs.delete_tds_by_tgId({'tg_id':request.form['tgs_delete']})
    Tgs.delete_tgs_by_tgId({'id':request.form['tgs_delete']})
    # Obtener circuitos y total_tds después de la eliminación
    index = Tgs.total_name_tg({'tg_id': request.form['tgs_delete']})
    index2 = Total_tds.name_total_tds({'tab_secondary': request.form['tgs_delete']})

    # Verificar si hay un único circuito y cumple la condición de tg_id con td_id=None
    if len(index) == 1 and (index2[0]['td_id'] is None or not index2[0]['td_id']):
        circuit_id = index[0]['id']
        Tgs.update_name({'id': circuit_id, 'name': 1})
        last_number = 1

    elif len(index2) == 1:  # Si solo hay un total_tds con tab_secondary
        circuit_id = index2[0]['id']
        Total_tds.update_name_total_td({'id': circuit_id, 'name': 1})
        last_number = 1

    else:
        # Ordenar y renombrar si hay más de un circuito
        ordered_circuits = sorted(index, key=lambda x: int(x['name']))
        new_names = list(range(1, len(ordered_circuits) + 1))
        updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}

        for circuit_id, new_name in updates.items():
            Tgs.update_name({'id': circuit_id, 'name': new_name})

        last_number = new_names[-1] if new_names else 0

        # Renombrar total_tds de manera correlativa
        new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
        updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}

        for circuit_id, new_name in updates_index2.items():
            Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})
    return redirect('/loadbox/')

@app.route('/api/delete/tds', methods=['POST'])
def delete_tds():
    td = Tds.tg_id_by_td_id({'id': request.form['tds_delete']})

    # Eliminar datos relacionados con el TD
    Tds.delete_load_by_tdId({'td_id': request.form['tds_delete']})
    Tds.delete_circuits_by_tdId({'td_id': request.form['tds_delete']})
    Tds.delete_total_tds_by_tdId({'td_id': request.form['tds_delete']})
    Tds.delete_tds_by_tdId({'id': request.form['tds_delete']})

    # Obtener circuitos y total_tds después de la eliminación
    index = Tgs.total_name_tg({'tg_id': td[0]['tg_id']})
    index2 = Total_tds.name_total_tds({'tab_secondary': td[0]['tg_id']})

    # Verificar si hay un único circuito y cumple la condición de tg_id con td_id=None
    if len(index) == 1 and (index2[0]['td_id'] is None or not index2[0]['td_id']):
        circuit_id = index[0]['id']
        Tgs.update_name({'id': circuit_id, 'name': 1})
        last_number = 1

    elif len(index2) == 1:  # Si solo hay un total_tds con tab_secondary
        circuit_id = index2[0]['id']
        Total_tds.update_name_total_td({'id': circuit_id, 'name': 1})
        last_number = 1

    else:
        # Ordenar y renombrar si hay más de un circuito
        ordered_circuits = sorted(index, key=lambda x: int(x['name']))
        new_names = list(range(1, len(ordered_circuits) + 1))
        updates = {circuit['id']: new_name for circuit, new_name in zip(ordered_circuits, new_names)}

        for circuit_id, new_name in updates.items():
            Tgs.update_name({'id': circuit_id, 'name': new_name})

        last_number = new_names[-1] if new_names else 0

        # Renombrar total_tds de manera correlativa
        new_names_index2 = list(range(last_number + 1, last_number + len(index2) + 1))
        updates_index2 = {circuit['id']: new_name for circuit, new_name in zip(index2, new_names_index2)}

        for circuit_id, new_name in updates_index2.items():
            Total_tds.update_name_total_td({'id': circuit_id, 'name': new_name})

    return redirect('/loadbox/')




@app.route('/api/delete/proyect', methods=['POST'])
def delete_pro():
    allpro = Proyect.get_all_tgs_by_proyect_id({'id':request.form['proyect_id']})
    if len(allpro) > 0:
        return jsonify(data=allpro, success=False)
    else:
        Proyect.delete_proyect_by_proyec_id({'id':request.form['proyect_id']})
        return jsonify(success=True)



# -------------------------------------------------------------------------------------------------------------------
# ------------------------------------SUMMARY PROYECTS --------------------------------------------------------------
# -------------------------------------------------------------------------------------------------------------------


@app.route('/api/excel_tds/<int:td_id>')
def tds_to_excel(td_id):
    data_list = Tds.detail_to_excel({'td_id': td_id})
    name_td = Tds.get_td_by_id({'id': td_id})

    def balancear_cargas(df):
        print("BALANCEANDO...")
        total_r = df['R [A]'].sum(skipna=True)
        total_s = df['S [A]'].sum(skipna=True)
        total_t = df['T [A]'].sum(skipna=True)

        fase_totales = {'R [A]': total_r, 'S [A]': total_s, 'T [A]': total_t}
        media = sum(fase_totales.values()) / 3
        fases_ordenadas = sorted(fase_totales, key=lambda x: fase_totales[x], reverse=True)
        print(media)
        iteraciones = 0
        while True:
            iteraciones += 1
            max_diff = max(abs(fase_totales[fases_ordenadas[0]] - media),
                        abs(fase_totales[fases_ordenadas[2]] - media))
            if max_diff < 0.05 * media:
                print(f"Desbalance aceptable alcanzado: {max_diff} (iteración {iteraciones})")
                break

            for fase_origen in fases_ordenadas:
                for fase_destino in fases_ordenadas[::-1]:
                    if fase_origen == fase_destino:
                        continue

                    diferencia = fase_totales[fase_origen] - fase_totales[fase_destino]
                    if diferencia < 0.02 * media:
                        continue

                    for i, row in df.iterrows():
                        if pd.notna(row[fase_origen]) and pd.isna(row[fase_destino]) and all(
                            pd.isna(row[f]) for f in ['R [A]', 'S [A]', 'T [A]'] if f != fase_origen
                        ):
                            valor_mover = row[fase_origen]
                            df.at[i, fase_destino] = valor_mover
                            df.at[i, fase_origen] = None

                            fase_totales[fase_origen] -= valor_mover
                            fase_totales[fase_destino] += valor_mover

                            fases_ordenadas = sorted(fase_totales, key=lambda x: fase_totales[x], reverse=True)
                            break

            if iteraciones > 10:
                print(f"Se alcanzó el límite de iteraciones (10) con un desbalance de {round(max_diff,2)}.")
                break
        return df

    if data_list:
        numeric_fields = [
            'Potencia por Carga [W]', 'Fp', 'Frecuencia [Hz]', 'Potencia Total [Kw]',
            'R [A]', 'S [A]', 'T [A]', 'Tensión [V]', 'Largo [m]', 'Vp [V]',
            'Conductor [mm2]', 'Canalización [mm]'
        ]
        data = []
        for tg in data_list:
            if tg['Cantidad'] is not None:
                tg['Cantidad'] = int(tg['Cantidad'])
                for field in numeric_fields:
                    if field in tg:
                        value = tg[field]
                        if isinstance(value, float):
                            tg[field] = float(value)
                        elif isinstance(value, str):
                            try:
                                tg[field] = float(value.replace(',', '.'))
                            except ValueError:
                                tg[field] = value
                data.append(tg)

        df = pd.DataFrame(data)

        desbalance_porcentaje = None

        if 380 not in df['Tensión [V]'].values:
            print("No hay trifásico, cargas monofasicas solamente")
            df['Intensidad [A]'] = df[['R [A]', 'S [A]', 'T [A]']].sum(axis=1)
            df = df.drop(columns=['R [A]', 'S [A]', 'T [A]'])
            cols = df.columns.tolist()
            pos = cols.index('Tensión [V]')
            cols.insert(pos, cols.pop(cols.index('Intensidad [A]')))
            df = df[cols]

            if 220 in df['Tensión [V]'].values:
                print("Voltaje 0.220 detectado: ajustando formato de Intensidad [A]...")
                df['Intensidad [A]'] = df['Intensidad [A]'].round(2)

            sum_values = {
                'Circuito': 'Total',
                'Cantidad': df['Cantidad'].sum(),
                'Potencia Total [Kw]': df['Potencia Total [Kw]'].sum(),
                'Intensidad [A]': df['Intensidad [A]'].sum()
            }
        else:
            has_trifasico = df[['R [A]', 'S [A]', 'T [A]']].notna().all(axis=1).any()
            if has_trifasico:
                print("Hay trifásico, se balancearán las cargas monofásicas...")
                df = balancear_cargas(df)

            total_r = df['R [A]'].sum(skipna=True)
            total_s = df['S [A]'].sum(skipna=True)
            total_t = df['T [A]'].sum(skipna=True)
            promedio = (total_r + total_s + total_t) / 3
            max_diff = max(abs(total_r - promedio), abs(total_s - promedio), abs(total_t - promedio))
            desbalance_porcentaje = round((max_diff / promedio) * 100, 2) if promedio else 0.0
            sum_values = {
                'Circuito': 'Total',
                'Cantidad': df['Cantidad'].sum(),
                'Potencia Total [Kw]': df['Potencia Total [Kw]'].sum(),
                'R [A]': total_r,
                'S [A]': total_s,
                'T [A]': total_t
            }

        sum_df = pd.DataFrame(sum_values, index=[0])
        df = pd.concat([df, sum_df], ignore_index=True)

        sheet_name = str(name_td[0]['name'])
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=3, startcol=1, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            for col in worksheet.iter_cols(min_row=4, min_col=2, max_row=worksheet.max_row):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length
                worksheet.column_dimensions[column].width = adjusted_width

            specific_columns = ['R [A]', 'S [A]', 'T [A]', 'Fp']
            for col_name in specific_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 2
                    col_letter = worksheet.cell(row=4, column=col_idx).column_letter
                    if col_name == 'Fp':
                        worksheet.column_dimensions[col_letter].width = 5
                    else:
                        worksheet.column_dimensions[col_letter].width = 7

            for row in worksheet.iter_rows(min_row=4, min_col=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            last_row = worksheet.max_row
            for cell in worksheet[last_row]:
                cell.font = Font(bold=True)

            thin = Side(border_style="thin", color="000000")
            for row in worksheet.iter_rows(min_row=4, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            for cell in worksheet[worksheet.max_row]:
                cell.font = Font(bold=True)

            if desbalance_porcentaje is not None:
                mensaje = f"Desbalance: {desbalance_porcentaje}%"
                fila_desbalance = worksheet.max_row + 1
                col_inicio = 2
                col_fin = worksheet.max_column

                worksheet.merge_cells(start_row=fila_desbalance, start_column=col_inicio,
                                    end_row=fila_desbalance, end_column=col_fin)

                celda = worksheet.cell(row=fila_desbalance, column=col_inicio)
                celda.value = mensaje
                celda.font = Font(bold=True, color="FF0000")
                celda.alignment = Alignment(horizontal='left', vertical='center')

                for col in range(col_inicio, col_fin + 1):
                    celda_borde = worksheet.cell(row=fila_desbalance, column=col)
                    celda_borde.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=Cuadro_de_Cargas_{name_td[0]["name"]}.xlsx'
        return response
    else:
        return "No hay circuitos, tablero de distribucíon vacío."


# -----------------------------------------------------

@app.route('/api/excel_tgs/<int:tg_id>')
def tgs_to_excel(tg_id):
    data_list = Tgs.detail_to_excel({'tg_id': tg_id})
    name_tg = Tgs.name_tg({'tg_id': tg_id})


    def balancear_cargas(df):
        print("BALANCEANDO...")
        total_r = df['R [A]'].sum(skipna=True)
        total_s = df['S [A]'].sum(skipna=True)
        total_t = df['T [A]'].sum(skipna=True)

        fase_totales = {'R [A]': total_r, 'S [A]': total_s, 'T [A]': total_t}
        media = sum(fase_totales.values()) / 3
        fases_ordenadas = sorted(fase_totales, key=lambda x: fase_totales[x], reverse=True)
        print(media)
        iteraciones = 0
        while True:
            iteraciones += 1
            max_diff = max(abs(fase_totales[fases_ordenadas[0]] - media),
                        abs(fase_totales[fases_ordenadas[2]] - media))
            if max_diff < 0.05 * media:
                print(f"Desbalance aceptable alcanzado: {max_diff} (iteración {iteraciones})")
                break

            for fase_origen in fases_ordenadas:
                for fase_destino in fases_ordenadas[::-1]:
                    if fase_origen == fase_destino:
                        continue

                    diferencia = fase_totales[fase_origen] - fase_totales[fase_destino]
                    if diferencia < 0.02 * media:
                        continue

                    for i, row in df.iterrows():
                        if pd.notna(row[fase_origen]) and pd.isna(row[fase_destino]) and all(
                            pd.isna(row[f]) for f in ['R [A]', 'S [A]', 'T [A]'] if f != fase_origen
                        ):
                            valor_mover = row[fase_origen]
                            df.at[i, fase_destino] = valor_mover
                            df.at[i, fase_origen] = None

                            fase_totales[fase_origen] -= valor_mover
                            fase_totales[fase_destino] += valor_mover

                            fases_ordenadas = sorted(fase_totales, key=lambda x: fase_totales[x], reverse=True)
                            break

            if iteraciones > 10:
                print(f"Se alcanzó el límite de iteraciones (10) con un desbalance de {round(max_diff,2)}.")
                break
        return df

    if data_list:
        numeric_fields = [
            'Potencia por Carga [W]', 'Fp', 'Frecuencia [Hz]', 'Potencia Total [Kw]',
            'R [A]', 'S [A]', 'T [A]', 'Tensión [V]', 'Largo [m]', 'Vp [V]',
            'Conductor [mm2]', 'Canalización [mm]'
        ]
        data = []
        for tg in data_list:
            if tg['Cantidad'] is not None:
                tg['Cantidad'] = int(tg['Cantidad'])
                for field in numeric_fields:
                    if field in tg:
                        value = tg[field]
                        if isinstance(value, float):
                            tg[field] = float(value)
                        elif isinstance(value, str):
                            try:
                                tg[field] = float(value.replace(',', '.'))
                            except ValueError:
                                tg[field] = value
                data.append(tg)

        df = pd.DataFrame(data)

        desbalance_porcentaje = None

        if 380 not in df['Tensión [V]'].values:
            print("No hay trifásico, cargas monofasicas solamente")
            df['Intensidad [A]'] = df[['R [A]', 'S [A]', 'T [A]']].sum(axis=1)
            df = df.drop(columns=['R [A]', 'S [A]', 'T [A]'])
            cols = df.columns.tolist()
            pos = cols.index('Tensión [V]')
            cols.insert(pos, cols.pop(cols.index('Intensidad [A]')))
            df = df[cols]

            if 220 in df['Tensión [V]'].values:
                print("Voltaje 0.220 detectado: ajustando formato de Intensidad [A]...")
                df['Intensidad [A]'] = df['Intensidad [A]'].round(2)

            sum_values = {
                'Circuito': 'Total',
                'Cantidad': df['Cantidad'].sum(),
                'Potencia Total [Kw]': df['Potencia Total [Kw]'].sum(),
                'Intensidad [A]': df['Intensidad [A]'].sum()
            }
        else:
            has_trifasico = df[['R [A]', 'S [A]', 'T [A]']].notna().all(axis=1).any()
            if has_trifasico:
                print("Hay trifásico, se balancearán las cargas monofásicas...")
                df = balancear_cargas(df)

            total_r = df['R [A]'].sum(skipna=True)
            total_s = df['S [A]'].sum(skipna=True)
            total_t = df['T [A]'].sum(skipna=True)
            promedio = (total_r + total_s + total_t) / 3
            max_diff = max(abs(total_r - promedio), abs(total_s - promedio), abs(total_t - promedio))
            desbalance_porcentaje = round((max_diff / promedio) * 100, 2) if promedio else 0.0
            sum_values = {
                'Circuito': 'Total',
                'Cantidad': df['Cantidad'].sum(),
                'Potencia Total [Kw]': df['Potencia Total [Kw]'].sum(),
                'R [A]': total_r,
                'S [A]': total_s,
                'T [A]': total_t
            }

        sum_df = pd.DataFrame(sum_values, index=[0])
        df = pd.concat([df, sum_df], ignore_index=True)

        sheet_name = str(name_tg[0]['name'])
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=3, startcol=1, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            for col in worksheet.iter_cols(min_row=4, min_col=2, max_row=worksheet.max_row):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length
                worksheet.column_dimensions[column].width = adjusted_width

            specific_columns = ['R [A]', 'S [A]', 'T [A]', 'Fp']
            for col_name in specific_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 2
                    col_letter = worksheet.cell(row=4, column=col_idx).column_letter
                    if col_name == 'Fp':
                        worksheet.column_dimensions[col_letter].width = 5
                    else:
                        worksheet.column_dimensions[col_letter].width = 7

            for row in worksheet.iter_rows(min_row=4, min_col=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            last_row = worksheet.max_row
            for cell in worksheet[last_row]:
                cell.font = Font(bold=True)

            thin = Side(border_style="thin", color="000000")
            for row in worksheet.iter_rows(min_row=4, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            for cell in worksheet[worksheet.max_row]:
                cell.font = Font(bold=True)

            if desbalance_porcentaje is not None:
                mensaje = f"Desbalance: {desbalance_porcentaje}%"
                fila_desbalance = worksheet.max_row + 1
                col_inicio = 2
                col_fin = worksheet.max_column

                worksheet.merge_cells(start_row=fila_desbalance, start_column=col_inicio,
                                    end_row=fila_desbalance, end_column=col_fin)

                celda = worksheet.cell(row=fila_desbalance, column=col_inicio)
                celda.value = mensaje
                celda.font = Font(bold=True, color="FF0000")
                celda.alignment = Alignment(horizontal='left', vertical='center')

                for col in range(col_inicio, col_fin + 1):
                    celda_borde = worksheet.cell(row=fila_desbalance, column=col)
                    celda_borde.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=Cuadro_de_Cargas_{name_tg[0]["name"]}.xlsx'
        return response
    else:
        return "No hay circuitos, tablero de distribucíon vacío.", 



#--------------------------------------------------------------------------------------------------------------------------------

@app.route('/api/summary-tds/', methods=["POST"])
def allcircuits_tds():
    detail = Tds.summary_circuits_tds({'td_id': request.form['id']})
    return jsonify(detail)


@app.route("/api/edit_tgs_name/", methods=["POST"])
def edit_tgs_name():

    if not Tgs.validate_edit(request.form):
        return redirect("/loadbox")
    else:
        tgsid = {
            'id': request.form['id'],
            'name': "Tablero " + request.form['name'],
            'tag': request.form['tag']
        }
        Tgs.edit_name(tgsid)
        flash("Datos editados correctamente", "success_edit_tg")
    return redirect('/loadbox')


@app.route("/api/edit_tds_name/", methods=["POST"])
def edit_tds_name():

    if not Tds.validate_edit(request.form):
        return redirect("/loadbox")    
    else:
        tdsid = {
            'id': request.form['id'],
            'name': "Tablero " + request.form['name'],
            'tag': request.form['tag']
        }
        Tds.edit_name(tdsid)
        Total_tds.edit_name_total({'ref': tdsid['name'], 'td_id':tdsid['id']})
        flash("Datos editados correctamente.", "success_edit_td")        
    return redirect('/loadbox')
